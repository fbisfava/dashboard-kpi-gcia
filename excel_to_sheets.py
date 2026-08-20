"""
Convierte el Excel local al formato que espera Google Sheets para el dashboard.
Uso: python excel_to_sheets.py
Genera: para_importar_sheets.csv (en la misma carpeta)
"""
import openpyxl
import csv
from datetime import datetime

EXCEL_PATH  = r"C:\Users\bistolfi.federico\Downloads\KPI's Gerencia Financiera ..xlsx"
OUTPUT_PATH = r"C:\Users\bistolfi.federico\Downloads\para_importar_sheets.csv"
SHEET_NAME  = "DATOS RIESGO Y RECUPERO"

# Mapeo: (nombre columna GSheets, columna Excel 1-indexed)
# El orden de esta lista define el orden de columnas en el CSV de salida.
COLUMNS = [
    # ── FECHA ──────────────────────────────────────────────────────────────────
    ("Año-Mes",                                    1),

    # ── COBRANZA (dashboard cols 1-6) ─────────────────────────────────────────
    ("Cobranza Mes Sin PP",                         2),
    ("Cobranza Mes Con PP",                         3),
    ("Cobranza del mes promedio",                   4),
    ("Cobranza 90 dias Sin PP",                     5),
    ("Cobranza 90 dias Con PP",                     6),
    ("Cobranza 90 dias promedio",                   7),

    # ── MOROSIDAD TC (dashboard cols 7-10) ────────────────────────────────────
    ("Morosidad 1-60d Montos TC",                   8),
    ("Morosidad 1-60d Q casos TC",                  9),
    ("Morosidad +60d Montos TC",                   10),
    ("Morosidad +60d Q casos TC",                  11),

    # ── MOROSIDAD PRÉSTAMOS (dashboard cols 11-14) ────────────────────────────
    ("Morosidad 1-60d Montos Prést.",              12),
    ("Morosidad 1-60d Q casos Prést.",             13),
    ("Morosidad +60d Montos Prést.",               14),
    ("Morosidad +60d Q casos Prést.",              15),

    # ── CUENTAS Y CARTERA (dashboard cols 15-21) ──────────────────────────────
    ("Cuentas inhabilitadas o DV",                 16),
    ("Cuentas habilitadas",                        17),
    ("Cuentas totales",                            18),
    ("Ratio IH sobre totales",                     19),
    ("Q clientes que pasan a AB",                  20),
    ("Monto que pasa a AB",                        21),
    ("Q de refinanciaciones",                      23),   # col 22 Excel es Score Veraz refin.

    # ── (cols 22-23 GSheets vacías — reservadas) ──────────────────────────────
    ("(reservado)",                               None),
    ("(reservado)",                               None),

    # ── SCORE VERAZ CARTERA (dashboard col 24) ────────────────────────────────
    ("Score Veraz promedio",                       47),

    # ── ROLL RATES (dashboard cols 25-30) ─────────────────────────────────────
    ("RR 1-30 Préstamos",                          48),
    ("RR 1-30 TC",                                 49),
    ("RR Directo 90-120d Préstamos",               50),
    ("RR Directo 90-120d TC",                      51),
    ("RR 1-30 Total",                              52),
    ("RR Directo 90-120d Total",                   53),

    # ── VINTAGE >90 (dashboard cols 31-32) ────────────────────────────────────
    ("Vintage >90 prést. a 6 meses",               56),
    ("Vintage >90 prést. a 12 meses",              57),

    # ── ORIGINACIÓN SIISA (dashboard cols 33-43) ──────────────────────────────
    ("N° Solicitantes General",                    58),
    ("Tasa Aprobación General",                    60),
    ("Tasa Rechazo General",                       61),
    ("N° Solicitantes Tarjeta",                    62),
    ("Tasa Aprobación Tarjeta",                    63),
    ("Tasa Rechazo Tarjeta",                       64),
    ("N° Solicitantes Préstamo",                   65),
    ("Tasa Aprobación Préstamo",                   66),
    ("Tasa Rechazo Préstamo",                      67),
    ("Rechazos Política Zonas Prést.",             68),
    ("Tasa de conversión Veraz",                   69),

    # ── NUEVOS — CUENTAS Y CARTERA (dashboard cols 44-51) ────────────────────
    ("Score Veraz promedio refinanciaciones",      22),
    ("FPD Refinanciaciones",                       24),
    ("% FPD refinanciaciones",                     25),
    ("Q préstamos",                                26),
    ("FPD préstamos",                              27),
    ("% FPD préstamos",                            28),
    ("Cuentas con préstamo activo",                38),
    ("% cuentas hab. con préstamo activo",         39),

    # ── NUEVOS — TASAS DE CURA (dashboard cols 52-60) ────────────────────────
    ("Tasa de cura préstamos T2",                  29),
    ("Tasa de cura TC T2",                         30),
    ("Tasa de cura refin. T2",                     31),
    ("Tasa de cura préstamos T3",                  32),
    ("Tasa de cura TC T3",                         33),
    ("Tasa de cura refin. T3",                     34),
    ("Tasa de cura préstamos T4",                  35),
    ("Tasa de cura TC T4",                         36),
    ("Tasa de cura refin. T4",                     37),

    # ── NUEVOS — ORIGINACIÓN ALTAS (dashboard cols 61-65) ────────────────────
    ("Cantidad de altas en el mes",                42),
    ("Altas sobre aprobados",                      43),
    ("Altas TC",                                   44),
    ("Altas SPP",                                  45),
    ("% altas TC con uso en primer mes",           46),

    # ── NUEVOS — VINTAGE >30 (dashboard cols 66-67) ──────────────────────────
    ("Vintage >30 prést. a 6 meses",               54),
    ("Vintage >30 prést. a 12 meses",              55),

    # ── RECUPERO — COMPOSICIÓN REFINANCIACIONES (dashboard cols 68-69) ───────
    ("Composición refi TC",                        70),
    ("Composición refi Préstamos",                 71),

    # ── RECUPERO — GESTIÓN Y COBRANZA (dashboard cols 70-78) ─────────────────
    ("Clientes en mora",                           72),
    ("Clientes en mora mes c gestion",             73),
    ("Clientes en mora c gestion positiva",        74),
    ("Cuentas con gestion x mes",                  75),
    ("Tasa de clientes en mora gestionados",       76),
    ("Tasa de cumplimiento de promesas",           77),
    ("Tasa de contacto efectivo",                  78),
    ("Tasa de conversion de gestion a pago",       79),
    ("Intensidad de gestion",                      80),
]


def fmt_date(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m")
    s = str(val)
    # Handle "2024-01-01 00:00:00" strings
    try:
        return datetime.strptime(s[:10], "%Y-%m-%d").strftime("%Y-%m")
    except ValueError:
        return s


def fmt_value(val):
    if val is None:
        return ""
    if isinstance(val, float):
        # Preserve full precision; Google Sheets handles it fine
        return repr(val)
    return str(val)


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    # Find last data row
    last_row = 2
    for r in range(3, 200):
        if ws.cell(row=r, column=1).value is not None:
            last_row = r
        else:
            break

    print(f"Leyendo filas 3 a {last_row} ({last_row - 2} períodos)")

    headers = [name for name, _ in COLUMNS]
    rows_out = [headers]

    for r in range(3, last_row + 1):
        row = []
        for name, excel_col in COLUMNS:
            if excel_col is None:
                row.append("")
                continue
            cell = ws.cell(row=r, column=excel_col)
            val  = cell.value
            if name == "Año-Mes":
                row.append(fmt_date(val))
            else:
                row.append(fmt_value(val))
        rows_out.append(row)

    with open(OUTPUT_PATH, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerows(rows_out)

    print(f"Archivo generado: {OUTPUT_PATH}")
    print(f"  {len(rows_out)-1} filas de datos, {len(headers)} columnas")
    print("\nPara importar en Google Sheets:")
    print("  Archivo → Importar → Subir → elegir 'para_importar_sheets.csv'")
    print("  Tipo de importación: 'Reemplazar hoja de cálculo actual'")
    print("  Separador: coma (detectado automáticamente)")


if __name__ == "__main__":
    main()
